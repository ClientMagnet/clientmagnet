from flask import Flask, request, jsonify, send_file, render_template, session
import requests
import re
import json
import urllib.parse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import time
import datetime
import os

app = Flask(__name__)
app.secret_key = "clientmagnet-secret-key-for-local-dev-12345"

def load_config():
    config_file = "config.json"
    if os.path.exists(config_file):
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading config: {e}")
    return {}

config = load_config()
GOOGLE_CLIENT_ID = config.get("GOOGLE_CLIENT_ID") or os.environ.get("GOOGLE_CLIENT_ID", "1008719970978-hb24n2q4kg77up0pl11674457ps7568c.apps.googleusercontent.com")
USERS_DB_FILE = "users.json"

def load_users():
    if not os.path.exists(USERS_DB_FILE):
        return {}
    try:
        with open(USERS_DB_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading users: {e}")
        return {}

def save_users(users):
    try:
        with open(USERS_DB_FILE, "w", encoding="utf-8") as f:
            json.dump(users, f, indent=4, ensure_ascii=False)
    except Exception as e:
        print(f"Error saving users: {e}")

def get_user_profile(email, name=None, picture=None):
    users = load_users()
    today_str = datetime.date.today().isoformat()
    
    if email not in users:
        users[email] = {
            "name": name or email.split("@")[0].capitalize(),
            "email": email,
            "picture": picture or "",
            "plan": "free",
            "leads_consumed": 15,
            "last_reset_date": today_str
        }
        save_users(users)
    else:
        user = users[email]
        # Reiniciar contador diario si la fecha cambió
        if user.get("last_reset_date") != today_str:
            user["leads_consumed"] = 0
            user["last_reset_date"] = today_str
            save_users(users)
            
        updated = False
        if name and user.get("name") != name:
            user["name"] = name
            updated = True
        if picture and user.get("picture") != picture:
            user["picture"] = picture
            updated = True
        if updated:
            save_users(users)
            
    return users[email]


HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "es-ES,es;q=0.9,en;q=0.8"
}

def clean_argentine_phone(raw_phone, default_area_code="351"):
    if not raw_phone:
        return None
        
    # Dejar solo dígitos
    digits = re.sub(r'\D', '', raw_phone)
    
    # Eliminar prefijo de país 54 si existe
    if digits.startswith("54"):
        digits = digits[2:]
        
    # Detectar y remover prefijo internacional de celulares '9' si está presente
    # E.g. en +54 9 11 234-5678, tras quitar 54 queda 9112345678 (10 dígitos)
    # E.g. en +54 9 351 15 234-5678, tras quitar 54 queda 9351152345678 (13 dígitos)
    has_international_mobile_9 = False
    if digits.startswith("9") and len(digits) in [10, 11, 12, 13]:
        digits = digits[1:]
        has_international_mobile_9 = True
        
    # Eliminar prefijo de salida local 0 si existe (ej. 0351 -> 351)
    if digits.startswith("0"):
        digits = digits[1:]
        
    # Si tiene 12 dígitos o si tiene 11 dígitos y no se quitó el 9 (lo que significa que podría tener el 15)
    # O si es len 11 y se quitó el 9 (ej. 9 + 351 + 15 + 6 digitos = 12 digitos, tras quitar 9 quedan 11 digitos)
    if len(digits) == 12 or (len(digits) == 11):
        # Detectar el "15" y removerlo según la longitud de la característica (área)
        if digits[2:4] == "15":
            digits = digits[:2] + digits[4:]
        elif digits[3:5] == "15":
            digits = digits[:3] + digits[5:]
        elif digits[4:6] == "15":
            digits = digits[:4] + digits[6:]
            
    # Si sólo tiene 7 dígitos (ej. 4823237), asumimos el código de área por defecto
    if len(digits) == 7:
        digits = default_area_code + digits
    # Si tiene 9 dígitos y empieza con 15 (ej. 158506067), removemos 15 y agregamos código
    elif len(digits) == 9 and digits.startswith("15"):
        digits = default_area_code + digits[2:]
        
    # Si ya tenemos 10 dígitos (ej. 3514823237), agregamos el código de WhatsApp 549
    if len(digits) == 10:
        return "549" + digits
        
    # Fallback final: Devolver 549 + los últimos 10 dígitos si es más largo
    if len(digits) > 10:
        return "549" + digits[-10:]
        
    return "549" + digits

def get_area_code_by_city(city):
    if not city:
        return "351"
    city_clean = city.lower()
    if "cordoba" in city_clean or "córdoba" in city_clean:
        return "351"
    elif "buenos" in city_clean or "caba" in city_clean or "capital" in city_clean:
        return "11"
    elif "rosario" in city_clean:
        return "341"
    elif "mendoza" in city_clean:
        return "261"
    elif "tucuman" in city_clean or "tucumán" in city_clean:
        return "381"
    elif "salta" in city_clean:
        return "387"
    elif "santa fe" in city_clean or "santa fé" in city_clean:
        return "342"
    elif "mar del plata" in city_clean:
        return "223"
    elif "bahia" in city_clean or "bahía" in city_clean:
        return "291"
    return "11"  # Por defecto Buenos Aires si no se reconoce

ALLOWED_SERVICES_KEYWORDS = [
    "web", "pagina", "página", "sitio", "desarrollo", "programador", "programacion", "programación",
    "software", "app", "aplicacion", "aplicación",
    "community", "redes", "social", "instagram", "facebook", "post", "contenido", "feed", "stories",
    "diseño", "diseno", "grafico", "gráfico", "branding", "identidad", "logo", "logotipo",
    "marketing", "publicidad", "ads", "anuncios", "seo", "google ads", "facebook ads", "tráfico", "trafico", "growth",
    "copywriting", "redaccion", "redacción", "copywriter", "copy",
    "empleado", "empleada", "trabajar", "puesto", "recepcion", "recepción", "recepcionista", "asistente",
    "secretaria", "secretario", "administra", "administración", "administracion", "soporte", "atencion", "atención",
    "ventas", "vendedor", "vendedora", "comercial", "setter", "closer", "telemarketer",
    "fotografia", "fotografía", "video", "edicion", "edición", "filmmaker", "camara", "cámara", "audiovisual"
]

def validate_service_string(service):
    if not service:
        return False
    service_clean = service.lower().strip()
    
    # Validar que el servicio ingresado contenga al menos una palabra clave profesional típica
    for keyword in ALLOWED_SERVICES_KEYWORDS:
        if keyword in service_clean:
            return True
    return False

import random
from concurrent.futures import ThreadPoolExecutor

def extract_all_external_urls(obj):
    found = set()
    if isinstance(obj, dict):
        for v in obj.values():
            found.update(extract_all_external_urls(v))
    elif isinstance(obj, list):
        for item in obj:
            found.update(extract_all_external_urls(item))
    elif isinstance(obj, str):
        if obj.startswith('http://') or obj.startswith('https://'):
            ignored = ["ggpht.com", "schema.org", "gstatic.com", "googleusercontent.com", 
                       "streetviewpixels", "googleapis.com", "w3.org", "local/reviews", "local/business"]
            if not any(x in obj for x in ignored):
                if "google.com" in obj and "sites.google.com" not in obj:
                    pass
                else:
                    found.add(obj)
    return found

def get_best_links(urls):
    primary_website = None
    instagram_url = None
    facebook_url = None
    booking_url = None
    whatsapp_url = None
    other_social_url = None
    
    sorted_urls = sorted(list(urls), key=len)
    
    for url in sorted_urls:
        url_lower = url.lower()
        if "instagram.com" in url_lower:
            instagram_url = url
        elif "facebook.com" in url_lower:
            facebook_url = url
        elif any(x in url_lower for x in ["wa.link", "wa.me", "api.whatsapp.com", "whatsapp.com"]):
            whatsapp_url = url
        elif any(x in url_lower for x in ["agendapro.com", "reserva", "booking", "turno", "simplero"]):
            booking_url = url
        elif any(x in url_lower for x in ["linktr.ee", "instabio", "linkbio"]):
            other_social_url = url
        else:
            if not any(x in url_lower for x in ["business.google.com", "google.com/business", "google.com/maps"]):
                primary_website = url

    if not primary_website:
        if instagram_url:
            primary_website = instagram_url
        elif facebook_url:
            primary_website = facebook_url
        elif other_social_url:
            primary_website = other_social_url
        elif whatsapp_url:
            primary_website = whatsapp_url
        elif booking_url:
            primary_website = booking_url

    return {
        "website": primary_website,
        "instagram": instagram_url,
        "facebook": facebook_url,
        "whatsapp": whatsapp_url,
        "booking": booking_url,
        "other_social": other_social_url
    }

def fallback_search_socials(business_name, city, platform="instagram"):
    # Limpiar caracteres especiales para evitar ruidos en la búsqueda
    clean_name = re.sub(r'[^\w\s]', '', business_name)
    query = f"{clean_name} {city} {platform}"
    
    # Intentar con DuckDuckGo primero
    ddg_url = f"https://html.duckduckgo.com/html/?q={urllib.parse.quote(query)}"
    try:
        resp = requests.get(ddg_url, headers=HEADERS, timeout=5)
        if resp.status_code == 200:
            html = resp.text
            pattern = r'https?://(?:www\.)?' + (r'instagram\.com' if platform == "instagram" else r'facebook\.com') + r'/[a-zA-Z0-9_\.\-]+'
            links = re.findall(pattern, html, re.IGNORECASE)
            filtered = []
            for l in links:
                l_lower = l.lower()
                if not any(x in l_lower for x in ["/p/", "/reel/", "/stories/", "/accounts/", "/explore/", "/developer", "/privacy", "/terms"]):
                    clean_link = re.sub(r'[.,;:\'"\)\\]+$', '', l)
                    filtered.append(clean_link)
            if filtered:
                return filtered[0]
    except Exception:
        pass
        
    # Intentar con Bing como respaldo
    bing_url = f"https://www.bing.com/search?q={urllib.parse.quote(query)}"
    try:
        resp = requests.get(bing_url, headers=HEADERS, timeout=5)
        if resp.status_code == 200:
            html = resp.text
            pattern = r'https?://(?:www\.)?' + (r'instagram\.com' if platform == "instagram" else r'facebook\.com') + r'/[a-zA-Z0-9_\.\-]+'
            links = re.findall(pattern, html, re.IGNORECASE)
            filtered = []
            for l in links:
                l_lower = l.lower()
                if not any(x in l_lower for x in ["/p/", "/reel/", "/stories/", "/accounts/", "/explore/", "/developer", "/privacy", "/terms"]):
                    clean_link = re.sub(r'[.,;:\'"\)\\]+$', '', l)
                    filtered.append(clean_link)
            if filtered:
                return filtered[0]
    except Exception:
        pass
        
    return None

def analyze_website_socials_and_cms(url):
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    result = {
        "instagram": None,
        "facebook": None,
        "builder": None,
        "status": "success",
        "latency": 0.0,
        "has_viewport": False,
        "has_ssl": False,
        "has_seo": False
    }
    
    if not url:
        result["status"] = "no_url"
        return result
        
    url_lower = url.lower()
    if any(x in url_lower for x in ["instagram.com", "facebook.com", "wa.link", "whatsapp.com", "linktr.ee", "youtube.com"]):
        result["instagram"] = url if "instagram.com" in url_lower else None
        result["facebook"] = url if "facebook.com" in url_lower else None
        result["status"] = "social_url"
        result["has_ssl"] = url_lower.startswith("https")
        return result
        
    try:
        target_url = url
        if not target_url.startswith("http"):
            target_url = "http://" + target_url
            
        start_time = time.perf_counter()
        
        # Descarga parcial de hasta 250KB con stream=True
        resp = requests.get(target_url, headers=HEADERS, timeout=3, verify=False, stream=True)
        
        result["latency"] = round(time.perf_counter() - start_time, 3)
        result["has_ssl"] = resp.url.lower().startswith("https")
        
        if resp.status_code != 200:
            result["status"] = f"http_{resp.status_code}"
            return result
            
        content_bytes = b""
        for chunk in resp.iter_content(chunk_size=10240):
            content_bytes += chunk
            if len(content_bytes) > 250 * 1024:
                break
                
        html = content_bytes.decode('utf-8', errors='ignore')
        html_lower = html.lower()
        
        ig_pattern = r'https?://(?:www\.)?instagram\.com/[a-zA-Z0-9_\.\-]+'
        fb_pattern = r'https?://(?:www\.)?facebook\.com/[a-zA-Z0-9_\.\-]+'
        
        ig_links = re.findall(ig_pattern, html, re.IGNORECASE)
        fb_links = re.findall(fb_pattern, html, re.IGNORECASE)
        
        ig_filtered = [l for l in ig_links if not any(x in l.lower() for x in ["/p/", "/reel/", "/stories/", "/accounts/", "/explore/", "/developer", "/privacy", "/terms"])]
        fb_filtered = [l for l in fb_links if not any(x in l.lower() for x in ["/sharer", "/plugins", "/tr", "/pages/create", "/groups", "/ads", "/privacy", "/terms"])]
        
        result["instagram"] = re.sub(r'[.,;:\'"\)\\]+$', '', ig_filtered[0]) if ig_filtered else None
        result["facebook"] = re.sub(r'[.,;:\'"\)\\]+$', '', fb_filtered[0]) if fb_filtered else None
        
        if "wp-content" in html_lower or "wp-includes" in html_lower or 'name="generator" content="wordpress"' in html_lower:
            result["builder"] = "WordPress"
        elif "wix.com" in html_lower or "wix-code" in html_lower or "wix-site" in html_lower or 'content="wix.com' in html_lower:
            result["builder"] = "Wix"
        elif "shopify.com" in html_lower or "cdn.shopify.com" in html_lower or "shopify.theme" in html_lower:
            result["builder"] = "Shopify"
        elif "squarespace.com" in html_lower or "static1.squarespace.com" in html_lower:
            result["builder"] = "Squarespace"
        elif "blogger.com" in html_lower or "blogspot.com" in html_lower:
            result["builder"] = "Blogger"
            
        if '<meta' in html_lower and 'name="viewport"' in html_lower:
            result["has_viewport"] = True
            
        has_title = "<title" in html_lower
        has_desc = 'name="description"' in html_lower or 'property="og:description"' in html_lower
        result["has_seo"] = has_title and has_desc
        
        return result
    except requests.exceptions.Timeout:
        result["status"] = "timeout"
    except Exception:
        result["status"] = "error"
        
    return result

def calculate_service_score(servicio, website, rating=None, instagram=None, facebook=None, builder=None, status="success", has_viewport=False, has_ssl=False, has_seo=False, latency=0.0):
    serv_lower = servicio.lower().strip()
    web_clean = website.lower() if website else ""
    
    # 1. Desarrollo Web
    if any(x in serv_lower for x in ["web", "pagina", "página", "sitio", "desarrollo"]):
        if not website:
            score = 15
            motivo = "Sin Web"
        else:
            if any(x in web_clean for x in ["instagram.com", "facebook.com", "wa.link", "whatsapp.com", "linktr.ee"]):
                score = 25
                motivo = "Solo Redes"
            elif status in ["timeout", "error"] or (isinstance(status, str) and status.startswith("http_")):
                score = 30
                motivo = "Web Caída"
            elif builder in ["Wix", "Blogger"] or any(x in web_clean for x in ["blogspot", "wix", "site123", "weebly"]):
                score = 55
                motivo = f"Web Amateur ({builder})" if builder else "Web Amateur"
            else:
                score = 80
                motivo = f"Web ({builder})" if builder else "Web Profesional"
                
            modifiers = 0
            if has_ssl:
                modifiers += 5
            if has_viewport:
                modifiers += 5
            if has_seo:
                modifiers += 5
            if latency > 0.0 and latency < 0.8:
                modifiers += 5
            elif latency > 0.0 and latency <= 2.0:
                modifiers += 3
                
            score = min(100, score + modifiers)
            
            if score < 100:
                carencias = []
                if not has_ssl: carencias.append("Sin SSL")
                if not has_viewport: carencias.append("No Mobile")
                if latency > 2.0: carencias.append("Lenta")
                if carencias:
                    motivo += " (" + ", ".join(carencias) + ")"
                    
        return score, motivo

    # 2. Community Manager / Redes
    elif any(x in serv_lower for x in ["community", "redes", "social", "instagram", "post", "diseño"]):
        inst = instagram
        face = facebook
        if not inst and "instagram.com" in web_clean:
            inst = website
        if not face and "facebook.com" in web_clean:
            face = website
            
        score = 0
        reasons = []
        if inst:
            score += 60
            reasons.append("Instagram")
        if face:
            score += 30
            reasons.append("Facebook")
        if website and not any(x in web_clean for x in ["instagram.com", "facebook.com"]):
            score += 10
            
        if score == 0:
            score = 15
            motivo = "Sin Redes"
        else:
            motivo = " + ".join(reasons) + " Detectado"
            
        return score, motivo

    # 3. Empleado / Asistente / Ventas / Otros
    else:
        if rating:
            score = int(float(rating) * 20)
            score = max(10, min(100, score))
            if score < 50:
                motivo = f"Reputación Baja ({rating}★)"
            elif score < 85:
                motivo = f"Reputación Regular ({rating}★)"
            else:
                motivo = f"Reputación Excelente ({rating}★)"
        else:
            score = 65
            motivo = "Reputación Promedio"
        return score, motivo

def generate_local_proposal(name, nicho, servicio, estilo_mensaje="argentino", ciudad=None, website=None, calidad_motivo=None, instagram=None, facebook=None):
    if estilo_mensaje == "hibrido":
        has_custom_web = website and not any(x in website.lower() for x in ["instagram.com", "facebook.com", "wa.link", "whatsapp.com", "linktr.ee"])
        zona_text = ciudad if ciudad else "su zona"
        
        if not has_custom_web:
            tiene_redes = instagram or facebook
            if tiene_redes:
                return (
                    f"Hola, ¿cómo estás? Vi tu negocio *{name}* analizando las opciones del rubro en {zona_text}.\n\n"
                    f"Noté que se manejan de forma muy activa con el Instagram, pero que todavía no cuentan con una página web propia. Hoy en día, depender solo de redes expone al negocio a que un cambio de algoritmo te baje las visitas, además de que se pierdan muchos clientes que buscan directo en Google y quieren ver un catálogo ordenado.\n\n"
                    f"Me dedico a diseñar páginas web pensadas exclusivamente para el rubro de {nicho}, enfocadas en que sea súper fácil e intuitivo para la gente de tu zona ver tus productos y hacer pedidos.\n\n"
                    f"Si te interesa conocer una propuesta visual o ver algunos ejemplos de lo que hago, avisame y lo conversamos sin compromiso. ¡Saludos!"
                )
            else:
                return (
                    f"Hola, ¿cómo estás? Vi tu negocio *{name}* analizando las opciones del rubro en {zona_text}.\n\n"
                    f"Noté que actualmente no tienen una presencia fuerte en internet ni cuentan con una página web propia. En este rubro, no figurar de forma clara digitalmente hace que se pierdan muchos clientes de la zona que buscan directo en Google y terminan yendo a la competencia que sí tiene un catálogo visible.\n\n"
                    f"Me dedico a diseñar páginas web pensadas exclusivamente para el rubro de {nicho}, enfocadas en que sea súper fácil e intuitivo para la gente de tu zona ver tus productos y hacer pedidos.\n\n"
                    f"Si te interesa conocer una propuesta visual o ver algunos ejemplos de lo que hago, avisame y lo conversamos sin compromiso. ¡Saludos!"
                )
        elif calidad_motivo and "indexar" in calidad_motivo.lower():
            return (
                f"Hola, ¿cómo estás? Estaba buscando opciones de {nicho} en {zona_text} y encontré la página web de *{name}*.\n\n"
                f"Te escribo de forma muy puntual porque noté un detalle: tu página actual no está vinculada a tu ficha de Google Maps. Esto significa que la gente que te busca en el mapa no tiene forma de hacer clic para ir a ver tus productos, y terminan yéndose con la competencia de la zona.\n\n"
                f"Me especializo en optimización digital, ayudando a que marcas de tu sector tengan todos sus canales conectados y visibles para no perder ventas.\n\n"
                f"Si te parece bien, te puedo mostrar en un minuto cómo corregir este enlace para potenciar las visitas de tu página. Quedo a disposición."
            )
        else:
            carencias = []
            if calidad_motivo:
                if "Sin SSL" in calidad_motivo: carencias.append("Google muestra un aviso que dice que el sitio no es seguro (lo que suele generar desconfianza en la gente a la hora de comprar)")
                if "No Mobile" in calidad_motivo: carencias.append("cuesta un poco leerla desde el teléfono sin tener que andar haciendo zoom")
                if "Lenta" in calidad_motivo: carencias.append("tarda bastante en cargar las imágenes")
            
            error_msg = "tarda bastante en cargar las imágenes y cuesta un poco leerla desde el teléfono sin tener que andar haciendo zoom"
            if carencias:
                error_msg = " y ".join(carencias)
                
            return (
                f"Hola, ¿cómo estás? Vi la página web de *{name}* navegando desde el celular y me pareció muy interesante tu propuesta en {zona_text}.\n\n"
                f"Te contacto porque noté que la página {error_msg}. Hoy más del 80% de las consultas de {nicho} se hacen desde el teléfono, y una web incómoda hace que los usuarios abandonen la navegación.\n\n"
                f"Trabajo en el rediseño y modernización de páginas web para negocios de {nicho}, logrando que sean súper rápidas, visualmente atractivas y cómodas para comprar desde cualquier celular.\n\n"
                f"Si estás con tiempo esta semana, te puedo compartir un par de sugerencias visuales para mejorar la experiencia de tus clientes, sin costo. ¡Saludos!"
            )

    nicho_lower = nicho.lower().strip()
    serv_lower = servicio.lower().strip()
    
    # 1. Determinar categoría de servicio
    if any(x in serv_lower for x in ["web", "pagina", "página", "sitio", "desarrollo", "programador", "programacion", "programación", "software", "app"]):
        serv_cat = "web"
    elif any(x in serv_lower for x in ["community", "redes", "social", "instagram", "facebook", "post", "contenido", "feed", "stories"]):
        serv_cat = "community"
    elif any(x in serv_lower for x in ["empleado", "empleada", "trabajar", "puesto", "recep", "asistente", "secretari", "administra", "soporte", "atencion", "atención"]):
        serv_cat = "empleado"
    elif any(x in serv_lower for x in ["ventas", "vendedor", "vendedora", "comercial", "setter", "closer", "telemarketer"]):
        serv_cat = "ventas"
    elif any(x in serv_lower for x in ["diseño", "diseno", "grafico", "gráfico", "branding", "identidad", "logo"]):
        serv_cat = "diseno"
    elif any(x in serv_lower for x in ["fotografia", "fotografía", "video", "edicion", "edición", "filmmaker", "camara", "cámara", "audiovisual"]):
        serv_cat = "audiovisual"
    elif any(x in serv_lower for x in ["marketing", "publicidad", "ads", "anuncios", "seo", "trafico", "tráfico"]):
        serv_cat = "marketing"
    else:
        serv_cat = "generic"

    # 2. Determinar categoría de nicho
    if any(x in nicho_lower for x in ["estetica", "estética", "spa", "beauty", "belleza", "laser", "láser", "cosmeto", "peluqu", "salon", "salón"]):
        nicho_cat = "estetica"
        sector = "el sector de la estética y el bienestar"
    elif any(x in nicho_lower for x in ["dental", "odontolog", "dentista", "ortodoncia", "clinica dental", "clínica dental"]):
        nicho_cat = "dental"
        sector = "el sector de la salud odontológica"
    elif any(x in nicho_lower for x in ["gimnasio", "gym", "fitness", "entrenamiento", "crossfit", "pilates", "yoga", "entrenador"]):
        nicho_cat = "gym"
        sector = "el sector del fitness y la salud"
    elif any(x in nicho_lower for x in ["consultorio", "medic", "médic", "clinica", "clínica", "salud", "pediatra", "kinesiolog", "kinesiólog", "doctor"]):
        nicho_cat = "salud"
        sector = "el sector de la salud"
    elif any(x in nicho_lower for x in ["restauran", "bar", "cafe", "café", "pizzer", "rotiser", "helader", "gastronom", "comida", "hamburgue", "parrilla"]):
        nicho_cat = "gastronomia"
        sector = "el sector gastronómico"
    elif any(x in nicho_lower for x in ["ropa", "vestimenta", "tienda", "local", "comercio", "boutique", "zapater", "retail", "fiambrer", "carnicer", "almacen", "almacén"]):
        nicho_cat = "comercio"
        sector = "el comercio y ventas de productos"
    else:
        nicho_cat = "generic"
        sector = f"el sector de {nicho}"

    # 3. Definir dolores comercialmente específicos (nicho x servicio)
    pain_points = {
        "web": {
            "estetica": "noté un detalle crítico en su perfil: dependen de enlaces externos genéricos o de mensajes directos para coordinar turnos. El dolor real es perder clientes nocturnos que quieren agendar fuera de hora y, al no recibir respuesta inmediata, se van con la competencia. Estás perdiendo facturación en piloto automático por no contar con una plataforma de reservas 24/7.",
            "dental": "noté que coordinar los turnos de manera manual por WhatsApp suele saturar la recepción y generar demoras. El dolor de esto es que los pacientes posponen la consulta si no ven la disponibilidad de turnos en tiempo real, lo que genera baches vacíos en la agenda de los profesionales que ya no se pueden recuperar.",
            "gym": "noté que no cuentan con un sitio web donde los alumnos puedan ver horarios, cupos y pagar cuotas. El dolor principal es el desorden administrativo de responder las mismas dudas de horarios por Instagram y el tiempo perdido controlando pagos a mano en planillas.",
            "salud": "noté que el agendamiento manual satura tu secretaría y causa demoras. El verdadero dolor es la pérdida de tiempo respondiendo preguntas repetitivas, la dificultad para reprogramar citas rápido y las cancelaciones sobre la hora que dejan baches imposibles de cubrir sin automatización.",
            "gastronomia": "noté que dependen de plataformas de delivery (PedidosYa/Rappi) o de mensajes directos. El gran dolor aquí es regalar del 15% al 30% de tu facturación en comisiones abusivas a terceros, y además no tener una base de datos propia de tus comensales para enviarles promociones.",
            "comercio": "noté que no cuentan con una tienda online auto-gestionable. El gran dolor de esto es depender del 'precio por privado' para cerrar cada venta. Si un cliente está interesado fuera del horario de atención y no le contestás rápido, la venta se pierde.",
            "generic": "noté un detalle clave: para el nivel que manejan, no tienen una web propia profesional. El verdadero dolor es depender únicamente de redes sociales que pueden cerrar o cambiar su algoritmo, además de perder la autoridad que una 'fachada digital' sólida les daría frente a competidores."
        },
        "community": {
            "estetica": "noté un dolor muy común en las redes de estética: al ser un rubro visual, no tener un feed estético impecable, no mostrar resultados del antes/después o publicar poco hace que tu negocio parezca inactivo o poco profesional, perdiendo clientes que buscan calidad y confianza.",
            "dental": "noté que las redes de odontología suelen publicar placas frías o muy técnicas. El dolor de esto es la falta de confianza; la gente tiene miedo al dentista, y no humanizar la clínica mostrando testimonios de pacientes reales o videos educativos del equipo hace que no consigas nuevos turnos.",
            "gym": "noté que en redes de gimnasios el mayor dolor es no transmitir la energía del lugar, la vibra de las clases o la comunidad activa. Si tu Instagram está en silencio, la gente no se motiva a ir y termina eligiendo gimnasios que sí muestran Reels dinámicos del ambiente diario.",
            "salud": "noté que en salud el mayor dolor en redes es la falta de contenido humano. Publicar solo placas de stock sin mostrar los rostros del equipo médico hace que la cuenta se sienta distante, perdiendo la oportunidad de educar al paciente y captar nuevas consultas.",
            "gastronomia": "noté que en gastronomía el mayor dolor en redes es el contenido poco tentador. Si tus fotos son oscuras, no subís Reels de la preparación del plato o no mostrás la experiencia del salón, tu marca no genera ese antojo inmediato que es el motor de las ventas.",
            "comercio": "noté que en comercios el dolor de las redes es el formato de 'catálogo aburrido'. Subir solo fotos estáticas de productos sin reels dinámicos mostrando ideas de uso o combinaciones hace que la interacción sea nula y dependas de rebajar precios para vender.",
            "generic": "noté un dolor muy frecuente en las redes comerciales: mantenerlas activas insume demasiado tiempo y muchas veces se publica sin una estrategia clara. Esto resulta en una cuenta silenciosa o posteos que solo consiguen likes vacíos pero no consultas de clientes reales."
        },
        "empleado": {
            "estetica": "sé que en estética la sobrecarga de estar atendiendo en camilla y al mismo tiempo tener que contestar WhatsApps de reservas y dudas es asfixiante. El dolor de esto es demorarte en contestar, lo que hace que los clientes se cansen de esperar y busquen otro centro.",
            "dental": "sé que la secretaría de una clínica dental se desborda respondiendo turnos, reprogramaciones y dudas administrativas. El dolor real es el estrés operativo y las demoras en las respuestas que deterioran la experiencia del paciente y te hacen perder consultas.",
            "gym": "sé que el día a día operativo en la recepción de un gimnasio es agotador. El dolor real es el caos de controlar accesos, cobrar cuotas y responder WhatsApps de interesados en horas pico, lo que te quita tiempo para brindar una buena atención personalizada.",
            "salud": "sé que administrar un consultorio médico insume demasiadas horas. El gran dolor de todo profesional de salud es tener que lidiar con la carga administrativa y la atención al cliente constante mientras atiende pacientes, provocando retrasos y un estrés operativo innecesario.",
            "gastronomia": "sé que el día a día operativo en gastronomía es muy demandante. El dolor real de estar cocinando o atendiendo el salón y, a la vez, responder consultas de WhatsApp o redes genera demoras, estrés y errores en las órdenes.",
            "comercio": "sé que en el comercio el día a día es muy demandante. El dolor real de estar atendiendo el local y, a la vez, responder consultas de WhatsApp genera demoras en las respuestas que ahuyentan a los clientes que quieren comprar ya.",
            "generic": "sé que a medida que un negocio crece, la carga diaria se vuelve asfixiante. El dolor de todo dueño es pasar más tiempo respondiendo mensajes, gestionando planillas y resolviendo trámites administrativos que concentrándose en el crecimiento del negocio."
        },
        "ventas": {
            "estetica": "noté que un dolor frecuente es que te entren consultas por redes de interesadas, pero se enfríen por no tener un seguimiento de ventas rápido y activo. El dolor es perder prospectos que ya mostraron interés pero no cerraron la cita por falta de un vendedor proactivo.",
            "dental": "noté que un gran dolor en las clínicas es la pérdida de pacientes que preguntan presupuestos por WhatsApp pero quedan en la nada. No contar con un proceso de seguimiento de ventas activo hace que más de la mitad de esos presupuestos nunca se concreten.",
            "gym": "noté que en el rubro del fitness es común que pregunten precios de pases y no se les vuelva a hablar. El dolor de esto es perder socios potenciales que necesitan un empujoncito comercial (un seguimiento con propuesta de valor) para decidirse a arrancar.",
            "salud": "noté que un dolor común es que las consultas de interesados por tratamientos médicos costosos se enfríen rápido. Sin un asesor de ventas que les dé un seguimiento consultivo y profesional, esos pacientes posponen la decisión de atenderse.",
            "gastronomia": "noté que en gastronomía el dolor clave es el seguimiento. Muchas personas preguntan disponibilidad para eventos o precios de catering y quedan en la nada por responder de forma pasiva sin iniciar una conversación de venta activa.",
            "comercio": "noté que muchas personas preguntan precios de productos y quedan en la nada. Responder de forma pasiva ('Hola, sale $X') sin iniciar una conversación de venta activa o enviar un mensaje de seguimiento hace que más del 70% de los interesados no compre.",
            "generic": "noté que muchas veces se invierte tiempo y dinero en atraer interesados, pero el gran dolor es que esos leads quedan en la nada por no contar con un proceso de ventas estructurado y proactivo que les dé seguimiento inmediato para concretar el cierre."
        },
        "diseno": {
            "estetica": "noté un desajuste entre el excelente nivel de tus tratamientos y su imagen visual en redes. El dolor real de usar plantillas genéricas de Canva o logos desactualizados es que bajan tu autoridad, impidiéndote cobrar lo que realmente vale tu servicio por parecer amateur.",
            "dental": "noté que la imagen visual de tu clínica se siente desactualizada. El dolor de esto es que en salud la prolijidad visual se asocia directamente con la higiene y profesionalismo médico; un branding descuidado ahuyenta a pacientes que buscan calidad.",
            "gym": "noté que la identidad visual de tu gimnasio y sus banners no transmiten la fuerza de la marca. El dolor de esto es que te hace ver como un gimnasio de barrio informal en lugar de un centro de entrenamiento premium, impidiéndote subir la cuota sin quejas.",
            "salud": "noté que la imagen institucional de tu consultorio carece de cohesión. El dolor de esto es que en salud la prolijidad visual genera confianza; una marca inconsistente o desactualizada hace que el paciente perciba un servicio de menor categoría.",
            "gastronomia": "noté que la identidad visual de su marca y sus cartas no son consistentes. El dolor real de un diseño descuidado es que reduce el deseo de compra. La gente asocia una imagen descuidada con un producto de baja calidad, prefiriendo pagar más en la competencia.",
            "comercio": "noté que la imagen de su comercio no tiene cohesión (logo, publicaciones, bolsas). El dolor real es que sin una marca atractiva, tus productos parecen genéricos, obligándote a competir únicamente por precio bajo en lugar de valor.",
            "generic": "noté que no cuentan con una identidad de marca unificada. El dolor comercial de esto es que sin un branding coherente y profesional, el negocio termina compitiendo únicamente por precio bajo en lugar de competir por el valor de su marca."
        },
        "audiovisual": {
            "estetica": "noté que las fotos de tus tratamientos tienen mala luz o se ven pixeladas. El dolor en estética es crítico: la falta de fotos profesionales del antes/después con nitidez hace que la gente dude de los resultados reales de tus tratamientos.",
            "dental": "noté que las fotos de tu clínica odontológica no transmiten su verdadera calidad. En salud, fotos oscuras o mal encuadradas no reflejan la higiene y la tecnología avanzada de tus consultorios, restando confianza al paciente.",
            "gym": "noté que las fotos de tus entrenamientos se ven oscuras o amateurs. El dolor en fitness es que la gente compra motivación y resultados; si no mostrás videos nítidos y dinámicos que contagien ganas de entrenar, no lográs enganchar socios nuevos.",
            "salud": "noté que el material visual del consultorio se ve amateur. El dolor es que en salud la imagen es todo; fotos de baja calidad no reflejan el profesionalismo médico ni el nivel de tus instalaciones, alejando a pacientes exigentes.",
            "gastronomia": "noté que sus fotos de platos no hacen justicia a su comida. El gran dolor en gastronomía es que la comida entra por los ojos; si las fotos son oscuras o planas, no tientan al cliente a pedir y se pierde el motor del deseo gastronómico.",
            "comercio": "noté que las fotos de sus productos se ven amateurs (fondos desordenados, sombras duras). El dolor comercial es que las imágenes de baja calidad devalúan tus productos, haciendo que duden de la calidad y busquen otras tiendas con mejor estética.",
            "generic": "noté que el material audiovisual de sus perfiles carece de calidad profesional. El verdadero dolor es que un video mal editado o fotos oscuras hacen que el negocio parezca un emprendimiento informal, alejando al público con mayor poder adquisitivo."
        },
        "marketing": {
            "estetica": "noté que su presencia digital depende netamente del alcance orgánico de Instagram, que hoy llega a muy poca gente. El dolor es el estancamiento: pasar meses sin caras nuevas en el centro por no tener anuncios dirigidos que capten clientas en tu zona exacta.",
            "dental": "noté que dependen enteramente del boca en boca para conseguir nuevos pacientes. El dolor de no hacer campañas activas de publicidad geolocalizada es que tu flujo de pacientes es impredecible y dependés de la suerte en lugar de tener un flujo constante.",
            "gym": "noté que su presencia digital depende del alcance orgánico del algoritmo. El dolor de esto es que estás perdiendo de captar a cientos de vecinos que viven a pocas cuadras, que buscan arrancar el gimnasio y que te elegirían si vieran un anuncio tuyo.",
            "salud": "noté que no están realizando campañas de publicidad en Google o redes. El dolor real es que cuando un paciente busca tu especialidad con urgencia en tu zona, termina yéndose con el médico que sí está pagando publicidad para aparecer primero.",
            "gastronomia": "noté que no hacen campañas activas de anuncios geolocalizados para captar clientes en las horas de almuerzo/cena. El dolor de depender del flujo de la calle o la suerte en Instagram es que tu facturación diaria se vuelve muy inestable.",
            "comercio": "noté que no están realizando anuncios dirigidos para captar clientes en su ciudad. El dolor de depender solo del alcance orgánico es que tus ventas fluctúan demasiado y te cuesta escalar la facturación de forma predecible.",
            "generic": "noté que dependen enteramente del alcance orgánico o el boca en boca para atraer nuevos clientes. El verdadero dolor de esto es la inestabilidad en la facturación y la falta de control sobre tu crecimiento, al no contar con un sistema predecible que inyecte nuevos interesados al negocio todos los días."
        },
        "generic": {
            "estetica": "noté un detalle clave: para el nivel que tienen en estética, hay aspectos en este área que se pueden profesionalizar. El dolor real es el tiempo que perdés en procesos manuales y la fuga de clientas que buscan una experiencia premium de principio a fin.",
            "dental": "noté que hay aspectos administrativos y de comunicación en su clínica que se pueden optimizar. El dolor de esto es el desorden operativo diario, que satura a los profesionales con tareas secundarias y desvía la atención del paciente.",
            "gym": "noté que en su gimnasio se pueden mejorar los procesos de atención y gestión de alumnos. El dolor real es la pérdida de socios por falta de seguimiento y el estrés de administrar los cupos y pagos de forma informal.",
            "salud": "noté que en su consultorio hay margen para profesionalizar procesos. El dolor comercial es el desorden diario que estresa al equipo de secretaría y genera retrasos en la atención médica, afectando la experiencia del paciente.",
            "gastronomia": "noté que en gastronomía hay detalles de gestión o comunicación que se pueden automatizar. El gran dolor de no optimizar esto es perder ventas diarias por lentitud operativa y sobrecargar al personal en las horas de mayor demanda.",
            "comercio": "noté que en su comercio hay procesos manuales que se pueden modernizar. El dolor real es el tiempo perdido atendiendo consultas repetitivas de stock o precios, lo que te quita foco de las tareas comerciales clave para crecer.",
            "generic": "noté que en {name} hay aspectos de este área que se pueden optimizar. El verdadero dolor de no contar con una estructura profesionalizada es el doble esfuerzo operativo que tenés que hacer a diario, cargándote de estrés y limitando tu facturación."
        }
    }

    # 4. Definir soluciones por tipo de servicio
    solutions = {
        "web": "una plataforma web impecable, minimalista y profesional para {name} donde tus clientes puedan ver tu catálogo completo, evacuar dudas frecuentes sin que tengas que tipear la misma respuesta 100 veces, y reservar turnos en piloto automático sincronizados con tu agenda 24/7.",
        "community": "estructurar una grilla estética y visualmente impecable para {name}, con contenidos estratégicos que eduquen sobre el valor de lo que hacés, y la configuración de anuncios dirigidos para que te entren consultas de clientes potenciales reales todos los días, sin que tengas que pasar horas editando.",
        "empleado": "sumarme a tu equipo en {name} para aliviar este dolor de raíz: encargarme de la gestión de turnos, responder mensajes de forma rápida en redes y dar soporte administrativo diario. De esta manera, delegás en mí la carga operativa pesada para que vos puedas concentrarte en el servicio y recuperar tu tranquilidad.",
        "ventas": "encargarme del seguimiento comercial y prospección activa en {name}: contactar de forma rápida y persuasiva a cada interesado que escriba, derribar sus dudas e incentivar el cierre para concretar la venta o reserva inmediatamente, aumentando tu tasa de conversión.",
        "diseno": "desarrollar una identidad de marca impecable para {name}: un logotipo profesional, una paleta de colores coherente y plantillas premium para redes sociales que eleven el valor percibido del negocio y transmitan máxima confianza.",
        "audiovisual": "crear contenido fotográfico y videos en formato Reels/TikTok profesionales para {name}: capturar tus instalaciones, productos o servicios con iluminación y encuadres cinematográficos que den una imagen premium, generen deseo y atraigan clientes de alto valor.",
        "marketing": "diseñar e implementar campañas de publicidad de pago (anuncios geolocalizados en Meta/Google) para {name}: posicionar tu negocio frente a los vecinos de tu zona exacta y generar un flujo predecible de consultas diarias listas para comprar.",
        "generic": "un plan de {servicio} impecable y profesional para {name} adaptado a tu identidad, diseñado para simplificar tus tareas diarias, elevar el valor percibido por tus clientes y permitirte escalar tu facturación con menos fricción."
    }

    pain = pain_points.get(serv_cat, pain_points["generic"]).get(nicho_cat, pain_points[serv_cat]["generic"])
    pain = pain.replace("{name}", name)
    
    sol = solutions.get(serv_cat, solutions["generic"])
    sol = sol.replace("{name}", name).replace("{servicio}", servicio)
    
    # 5. Llamada a la acción / Cierre por tipo de servicio y estilo
    if estilo_mensaje == "corporativo":
        # Propuesta en Español Neutro Corporativo y Estético
        if serv_cat == "web":
            pain = f"notamos que actualmente se enfocan en un perfil comercial en redes sociales. Aunque es excelente para interacción rápida, consolidar un sitio web propio les permitiría construir un *ecosistema digital seguro* que resguarde su marca de los cambios de algoritmos y posicione a *{name}* como el referente indiscutido de su sector."
            sol = f"una plataforma web impecable, minimalista y profesional para *{name}* donde tus clientes puedan ver tu catálogo completo, evacuar dudas frecuentes de forma estética y automatizar la captación de reservas las 24 horas."
            cta = "He diseñado un boceto de interfaz interactiva y optimizada para su marca. ¿Estarían abiertos a conocer esta propuesta visual sin ningún compromiso?"
        elif serv_cat == "community":
            pain = f"notamos que sus perfiles comerciales requieren mayor cohesión estética para transmitir la solidez de su marca. Consolidar una identidad visual robusta les permitirá destacar y capturar la atención de su público objetivo en su mercado local."
            sol = f"diseñar una grilla visual premium y estructurar campañas de anuncios dirigidos para posicionar a *{name}* como la opción preferente de su nicho."
            cta = "He preparado un boceto rápido de grilla de contenidos de ejemplo para ustedes. ¿Estarían abiertos a que les comparta una captura de pantalla para evaluar la dirección visual?"
        else:
            pain = f"notamos margen para profesionalizar y automatizar los procesos de comunicación en su área. Centralizar sus gestiones les permitiría liberar tiempo operativo y elevar el nivel de atención."
            sol = f"un plan de servicios de *{servicio}* adaptado a la identidad de *{name}*, orientado a optimizar la eficiencia y elevar la autoridad de la marca."
            cta = "He preparado una breve propuesta estructurada de optimización. ¿Estarían abiertos a conocerla sin ningún compromiso?"
            
        proposal = (
            f"Hola, equipo de *{name}*.\n\n"
            f"Analizando el posicionamiento digital de las marcas de {nicho} en su localidad, hemos estado observando su presencia en línea. "
            f"Con respecto al nivel de profesionalismo que manejan, {pain}\n\n"
            f"Por esta razón, consideramos que les aportaría gran valor implementar {sol}\n\n"
            f"{cta}\n\n"
            f"Saludos cordiales,\nFrancisco."
        )
    else:
        # Propuesta en Estilo Argentino Cercano y Voseo
        if serv_cat == "web":
            cta = "Armé un boceto visual rápido de cómo se vería su sitio renovado y optimizado. Si les interesa ver la captura o que les grabe un videito de 40 segundos, ¿están abiertos a que se los mande?"
        elif serv_cat == "community":
            cta = "Armé una propuesta visual rápida de grilla estética y anuncios para ustedes. Si les interesa ver una captura o un videito corto explicativo, ¿están abiertos a que se los comparta?"
        elif serv_cat == "empleado":
            cta = "Si les interesa conocer más sobre mi perfil y ver cómo podría facilitar su día a día, ¿están abiertos a que les comparta mi CV y propuesta sin compromiso?"
        elif serv_cat == "ventas":
            cta = "Armé un esquema rápido de seguimiento de ventas específico para ustedes. ¿Están abiertos a que les mande un audio corto de 1 minuto explicándolo?"
        elif serv_cat == "diseno":
            cta = "Armé una propuesta de rediseño de marca visual rápida para ustedes. ¿Están abiertos a que les envíe el boceto interactivo para ver qué les parece?"
        elif serv_cat == "audiovisual":
            cta = "Diseñé unas ideas visuales rápidas de contenido para su rubro. ¿Están abiertos a que les comparta un boceto de grilla de Reels de ejemplo?"
        elif serv_cat == "marketing":
            cta = "Armé una estructura rápida de campaña geolocalizada para su zona. ¿Están abiertos a que les mande un audio rápido explicándolo sin compromiso?"
        else:
            cta = "Armé una propuesta visual rápida sobre esto para ustedes. ¿Están abiertos a que les comparta una captura de pantalla del boceto para ver qué les parece?"

        proposal = (
            f"¡Buenas! ¿Cómo va?\n\n"
            f"Vi que la gente de {name} la está rompiendo con su propuesta en {sector}, "
            f"pero noté un detalle clave: {pain}\n\n"
            f"Por eso pensé que les vendría genial {sol}\n\n"
            f"{cta}\n\n"
            f"Saludos, Francisco."
        )
    
    return proposal

def generate_with_gemini(api_key, name, nicho, ciudad, estado_digital, servicio, template_body, estilo_mensaje="argentino"):
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={api_key}"
    headers = {"Content-Type": "application/json"}
    
    if estilo_mensaje == "corporativo":
        prompt = f"""
        Eres un Agente de Inteligencia Artificial experto en validación de datos comerciales y Copywriting B2B de alto rendimiento. Tu objetivo es auditar la información recopilada por el scraper, verificar la veracidad de su estado digital mediante búsquedas en tiempo real y, posteriormente, redactar una propuesta ultra-personalizada.

        Utiliza tu herramienta de búsqueda integrada (Google Search) para contrastar los datos antes de emitir un juicio en el mensaje.

        PROCESO DE AUDITORÍA CRÍTICA (Paso Obligatorio antes de redactar):
        1. Analiza los datos de entrada del backend:
           - Nombre del Negocio: {name}
           - Ciudad/Zona: {ciudad}
           - Nicho: {nicho}
           - Estado según Scraper: {estado_digital} (Ej: "Sin Web" o "Solo Redes")

        2. Ejecuta una búsqueda de contra-verificación en Google Search:
           - Haz una consulta en tiempo real utilizando la estructura: "{name} {nicho} {ciudad} sitio web" o "{name} oficial".
           - Analiza los primeros resultados orgánicos.
           - CRITERIO DE VERDAD: Si encuentras un dominio propio activo que corresponda inequívocamente al negocio (ej: supropiaweb.com o supropiaweb.tiendanube.com), ignora la etiqueta "{estado_digital}" del scraper si ésta decía "Sin Web". Si los resultados solo muestran perfiles de Instagram, Facebook, Páginas Amarillas o la propia ficha de Maps, confirma que efectivamente NO tienen sitio web propio.

        3. Determina el "Dolor Real":
           - CASO A (Confirmado SIN WEB): Tu ángulo será la ausencia de un canal propio, la dependencia de algoritmos de terceros y la pérdida de clientes locales que buscan en Google.
           - CASO B (El scraper dijo "Sin Web" pero tu búsqueda ENCONTRÓ WEB): Tu ángulo cambia drásticamente. El dolor ya no es "no tener web", sino un problema de visibilidad o SEO (ej: "Tu sitio web no está indexado correctamente en tu ficha de Maps, lo que hace que muchos clientes potenciales no lo encuentren al buscar en el mapa").
           - CASO C (Tiene web pero el scraper detectó fallas técnicas o carencias como {estado_digital}): Usa los datos de rendimiento/carencias (Falta de SSL, no es apto para celulares/no mobile, o es lento).

        REGLAS DE REDACCIÓN Y ESTÉTICA (Estilo Profesional / Neutro):
        - Redacta en un español neutro, sofisticado y corporativo (Usa el "tú" o "ustedes" de forma sutil).
        - Estructura visualmente limpia para WhatsApp usando saltos de línea y negritas (`*texto*`) para conceptos clave. Máximo 1 emoji sobrio.
        - El mensaje debe ser directo, omitiendo introducciones genéricas automatizadas.
        - Firma al final EXACTAMENTE como "Saludos cordiales,\nFrancisco." (sin añadidos de IA ni textos introductorios). El resultado debe ser directamente el mensaje listo para copiar y enviar.
        - Devuelve ÚNICAMENTE la propuesta final redactada en español, sin comentarios extras.
        """
    elif estilo_mensaje == "hibrido":
        prompt = f"""
        Eres un Consultor de Crecimiento Digital y Copywriter B2B. Tu objetivo es redactar un mensaje de primer contacto vía WhatsApp para ofrecer servicios de Diseño y Desarrollo Web. El tono debe ser una combinación exacta, sobria y profesional entre el español neutro y el argentino cercano.

        PROHIBICIONES ESTRICTAS:
        1. SIN TECNICISMOS: Prohibido usar términos como SSL, responsivo, indexación, hosting, etc. Habla en el idioma del comerciante (ventas, comodidad visual, confianza, velocidad).
        2. REGLA DE HONESTIDAD RADICAL (PROHIBIDO MENTIR O ADULAR): Está estrictamente prohibido usar frases prefabricadas como "Vi que tienen una gran presencia en redes" o "Me encantó tu contenido" si no es real o si el negocio no tiene redes activas. Debes basar tu comentario estrictamente en la realidad encontrada en tu búsqueda. Si no encuentras redes o están abandonadas, dilo de forma profesional o no lo menciones.

        Utiliza tu herramienta de búsqueda integrada (Google Search) para contrastar los datos antes de emitir un juicio en el mensaje.

        PROCESO DE AUDITORÍA CRÍTICA (Paso Obligatorio antes de redactar):
        1. Analiza los datos de entrada del backend:
           - Nombre del Negocio: {name}
           - Ciudad/Zona: {ciudad}
           - Nicho: {nicho}
           - Estado según Scraper: {estado_digital} (Ej: "Sin Web" o "Solo Redes")

        2. Ejecuta una búsqueda de contra-verificación en Google Search:
           - Haz una consulta en tiempo real utilizando la estructura: "{name} {nicho} {ciudad} sitio web" o "{name} oficial".
           - Analiza los primeros resultados orgánicos y sus redes.
           - CRITERIO DE VERDAD: Si encuentras un dominio propio activo que corresponda inequívocamente al negocio (ej: supropiaweb.com o supropiaweb.tiendanube.com), ignora la etiqueta "{estado_digital}" del scraper si ésta decía "Sin Web". Si los resultados solo muestran perfiles de Instagram, Facebook, Páginas Amarillas o la propia ficha de Maps, confirma que efectivamente NO tienen sitio web propio.

        3. Determina el "Dolor Real" y Clasifica la Redacción:
           - CASO A (Confirmado SIN WEB): Tu ángulo será la ausencia de un canal propio, la dependencia de algoritmos de terceros y la pérdida de clientes locales. Analiza si tienen redes sociales activas o no.
             * Si tienen redes activas y cuidadas, usa la plantilla "SI TIENE BUENAS REDES".
             * Si no tienen redes o están abandonadas, usa la plantilla "SI NO TIENE REDES O ESTÁN ABANDONADAS".
           - CASO B (El scraper dijo "Sin Web" pero tu búsqueda ENCONTRÓ WEB): Tu ángulo cambia drásticamente. El dolor ya no es "no tener web", sino que la página está escondida en la ficha de Google Maps, haciéndele perder clientes de la zona.
           - CASO C (Tiene web pero con fallas técnicas o carencias como {estado_digital}): Enfócate en la frustración del usuario común: si la página tarda bastante en cargar las imágenes, cuesta leerla desde el teléfono (sin viewport) o si muestra un aviso de que el sitio no es seguro (SSL).

        REGLAS DE TONO Y ESTILO (HÍBRIDO SOBRIO):
        - Usa el "voseo" de forma muy sutil y natural (ej: "cómo estás", "vi tu negocio", "si te interesa"). Está prohibido usar palabras lunfardas o exageradas como "che", "buenísimo", "bárbaro" o "pinta".
        - Mantén la elegancia y el respeto de la comunicación corporativa, pero con la cercanía de una conversación uno a uno en WhatsApp.
        - Estructura visual: Mensajes cortos, limpios, con saltos de línea y uso quirúrgico de negritas (`*texto*`) para destacar el nombre del negocio o el beneficio. Máximo 1 emoji profesional (ej: 💼 o 📈).
        - Firma al final con el cierre y firma correspondientes a los ejemplos provistos.
        - Devuelve ÚNICAMENTE la propuesta final redactada en español, sin comentarios extras.

        Sigue estrictamente las siguientes plantillas de redacción para cada escenario:

        ---
        ### ESCENARIO A: El negocio efectivamente NO tiene sitio web (Confirmado en la búsqueda)
        
        [PLANTILLA - SI TIENE BUENAS REDES]:
        Hola, ¿cómo estás? Vi tu negocio *{name}* analizando las opciones del rubro en {ciudad}.
        Noté que se manejan de forma muy activa con el Instagram, pero que todavía no cuentan con una página web propia. Hoy en día, depender solo de redes expone al negocio a que un cambio de algoritmo te baje las visitas, además de que se pierdan muchos clientes que buscan directo en Google y quieren ver un catálogo ordenado.
        
        [PLANTILLA - SI NO TIENE REDES O ESTÁN ABANDONADAS]:
        Hola, ¿cómo estás? Vi tu negocio *{name}* analizando las opciones del rubro en {ciudad}.
        Noté que actualmente no tienen una presencia fuerte en internet ni cuentan con una página web propia. En este rubro, no figurar de forma clara digitalmente hace que se pierdan muchos clientes de la zona que buscan directo en Google y terminan yendo a la competencia que sí tiene un catálogo visible.

        Me dedico a diseñar páginas web pensadas exclusivamente para el rubro de {nicho}, enfocadas en que sea súper fácil e intuitivo para la gente de tu zona ver tus productos y hacer pedidos.
        Si te interesa conocer una propuesta visual o ver algunos ejemplos de lo que hago, avisame y lo conversamos sin compromiso. ¡Saludos!

        ---
        ### ESCENARIO B: El scraper decía "Sin Web" pero tu búsqueda ENCONTRÓ un sitio web activo en Google
        Hola, ¿cómo estás? Estaba buscando opciones de {nicho} en {ciudad} y encontré la página web de *{name}*. 
        Te escribo de forma muy puntual porque noté un detalle: tu página actual no está vinculada a tu ficha de Google Maps. Esto significa que la gente que te busca en el mapa no tiene forma de hacer clic para ir a ver tus productos, y terminan yéndose con la competencia de la zona.
        Me especializo en optimización digital, ayudando a que marcas de tu sector tengan todos sus canales conectados y visibles para no perder ventas.
        Si te parece bien, te puedo mostrar en un minuto cómo corregir este enlace para potenciar las visitas de tu página. Quedo a disposición.

        ---
        ### ESCENARIO C: El negocio TIENE sitio web pero cuenta con fallas técnicas (Lento, sin SSL o no apto para celular)
        Hola, ¿cómo estás? Vi la página web de *{name}* navegando desde el celular y me pareció muy interesante tu propuesta en {ciudad}.
        Te contacto porque noté que la página experimenta algunas dificultades (como demoras en la carga de imágenes, cuesta leerla desde el teléfono o muestra un aviso de que no es segura: {estado_digital}). Hoy más del 80% de las consultas de {nicho} se hacen desde el teléfono, y una web incómoda hace que los usuarios abandonen la navegación.
        Trabajo en el rediseño y modernización de páginas web para negocios de {nicho}, logrando que sean súper rápidas, visualmente atractivas y cómodas para comprar desde cualquier celular.
        Si estás con tiempo esta semana, te puedo compartir un par de sugerencias visuales para mejorar la experiencia de tus clientes, sin costo. ¡Saludos!
        """
    else:
        prompt = f"""
        Eres un Agente de Inteligencia Artificial experto en validación de datos comerciales y Copywriting B2B de alto rendimiento. Tu objetivo es auditar la información recopilada por el scraper, verificar la veracidad de su estado digital mediante búsquedas en tiempo real y, posteriormente, redactar una propuesta ultra-personalizada.

        Utiliza tu herramienta de búsqueda integrada (Google Search) para contrastar los datos antes de emitir un juicio en el mensaje.

        PROCESO DE AUDITORÍA CRÍTICA (Paso Obligatorio antes de redactar):
        1. Analiza los datos de entrada del backend:
           - Nombre del Negocio: {name}
           - Ciudad/Zona: {ciudad}
           - Nicho: {nicho}
           - Estado según Scraper: {estado_digital} (Ej: "Sin Web" o "Solo Redes")

        2. Ejecuta una búsqueda de contra-verificación en Google Search:
           - Haz una consulta en tiempo real utilizando la estructura: "{name} {nicho} {ciudad} sitio web" o "{name} oficial".
           - Analiza los primeros resultados orgánicos.
           - CRITERIO DE VERDAD: Si encuentras un dominio propio activo que corresponda inequívocamente al negocio (ej: supropiaweb.com o supropiaweb.tiendanube.com), ignora la etiqueta "{estado_digital}" del scraper si ésta decía "Sin Web". Si los resultados solo muestran perfiles de Instagram, Facebook, Páginas Amarillas o la propia ficha de Maps, confirma que efectivamente NO tienen sitio web propio.

        3. Determina el "Dolor Real":
           - CASO A (Confirmado SIN WEB): Tu ángulo será la ausencia de un canal propio, la dependencia de algoritmos de terceros y la pérdida de clientes locales que buscan en Google.
           - CASO B (El scraper dijo "Sin Web" pero tu búsqueda ENCONTRÓ WEB): Tu ángulo cambia drásticamente. El dolor ya no es "no tener web", sino un problema de visibilidad o SEO (ej: "Tu sitio web no está indexado correctamente en tu ficha de Maps, lo que hace que muchos clientes potenciales no lo encuentren al buscar en el mapa").
           - CASO C (Tiene web pero el scraper detectó fallas técnicas o carencias como {estado_digital}): Usa los datos de rendimiento/carencias (Falta de SSL, no es apto para celulares/no mobile, o es lento).

        REGLAS ESTRICTAS DE REDACCIÓN Y ESTILO (Tono Argentino / Cercano):
        - TONO: Argentino, casual, profesional pero muy cercano. Usa "voseo" de forma natural (ej: "cómo estás", "vi tu perfil", "pensaba", "tenés"). Nada de "Estimado" ni formalidades rígidas de venta.
        - GANCHO INICIAL (Primeras 2 líneas): Prohibido usar "Estuve viendo tu perfil y me pareció excelente..." o "Te contacto porque...". Empieza directo al grano mencionando el nombre específico del negocio ({name}) de forma orgánica.
        - Estructura visualmente limpia para WhatsApp usando saltos de línea y negritas (`*texto*`) para conceptos clave. Máximo 1 emoji sobrio.
        - Firma al final EXACTAMENTE como "Saludos, Francisco." (sin añadidos de IA ni textos introductorios). El resultado debe ser directamente el mensaje listo para copiar y enviar.
        - Devuelve ÚNICAMENTE la propuesta final redactada en español, sin comentarios extras.
        """
    
    payload = {
        "contents": [{
            "parts": [{
                "text": prompt
            }]
        }],
        "tools": [{"google_search": {}}]
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=12)
        if response.status_code == 200:
            res_data = response.json()
            text = res_data['candidates'][0]['content']['parts'][0]['text'].strip()
            return text
    except Exception as e:
        print(f"Error llamando a Gemini API: {e}")
    return None

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/login")
def login_page():
    config = load_config()
    client_id = config.get("GOOGLE_CLIENT_ID") or os.environ.get("GOOGLE_CLIENT_ID", "1008719970978-hb24n2q4kg77up0pl11674457ps7568c.apps.googleusercontent.com")
    return render_template("login.html", google_client_id=client_id)

@app.route("/api/auth/google", methods=["POST"])
def auth_google():
    data = request.json or {}
    token = data.get("credential")
    if not token:
        return jsonify({"status": "error", "message": "Falta la credencial de Google."}), 400
        
    try:
        # Validar el token con la API de Google de manera segura y sin instalar librerías extra
        resp = requests.get(f"https://oauth2.googleapis.com/tokeninfo?id_token={token}", timeout=5)
        if resp.status_code != 200:
            return jsonify({"status": "error", "message": "Token de Google inválido o expirado."}), 401
            
        payload = resp.json()
        email = payload.get("email")
        name = payload.get("name", email.split("@")[0].capitalize())
        picture = payload.get("picture", "")
        
        if not email:
            return jsonify({"status": "error", "message": "El token de Google no contiene correo electrónico."}), 400
            
        # Obtener o registrar perfil del usuario
        user_profile = get_user_profile(email, name, picture)
        
        # Guardar datos básicos en la sesión de Flask
        session["user"] = {
            "email": email,
            "name": name,
            "picture": picture
        }
        
        return jsonify({"status": "success", "user": user_profile})
        
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error de autenticación: {str(e)}"}), 500

@app.route("/api/auth/google/redirect", methods=["POST"])
def auth_google_redirect():
    token = request.form.get("credential")
    if not token:
        return redirect("/login?error=no_credential")
        
    try:
        resp = requests.get(f"https://oauth2.googleapis.com/tokeninfo?id_token={token}", timeout=5)
        if resp.status_code != 200:
            return redirect("/login?error=invalid_token")
            
        payload = resp.json()
        email = payload.get("email")
        name = payload.get("name", email.split("@")[0].capitalize())
        picture = payload.get("picture", "")
        
        if not email:
            return redirect("/login?error=no_email")
            
        user_profile = get_user_profile(email, name, picture)
        
        session["user"] = {
            "email": email,
            "name": name,
            "picture": picture
        }
        
        return redirect("/")
        
    except Exception as e:
        print(f"Redirect Auth Error: {e}")
        return redirect("/login?error=server_error")

@app.route("/api/auth/demo", methods=["POST"])
def auth_demo():
    email = "demo.developer@clientmagnet.com"
    name = "Dev ClientMagnet"
    picture = ""
    
    user_profile = get_user_profile(email, name, picture)
    
    session["user"] = {
        "email": email,
        "name": name,
        "picture": picture
    }
    
    return jsonify({"status": "success", "user": user_profile})

@app.route("/api/auth/session", methods=["GET"])
def auth_session():
    user_session = session.get("user")
    if not user_session:
        return jsonify({"status": "error", "message": "No hay sesión activa."}), 401
        
    email = user_session["email"]
    user_profile = get_user_profile(email)
    
    return jsonify({"status": "success", "user": user_profile})

@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    session.pop("user", None)
    return jsonify({"status": "success", "message": "Sesión cerrada correctamente."})

@app.route("/api/upgrade", methods=["POST"])
def upgrade_plan():
    user_session = session.get("user")
    if not user_session:
        return jsonify({"status": "error", "message": "Debe iniciar sesión para realizar esta acción."}), 401
        
    email = user_session["email"]
    users = load_users()
    if email in users:
        users[email]["plan"] = "premium"
        save_users(users)
        return jsonify({"status": "success", "user": users[email]})
    else:
        return jsonify({"status": "error", "message": "Usuario no encontrado."}), 404


def process_candidate_details(c, ciudad, servicio, nicho, api_key, estilo_mensaje="argentino"):
    web = c["website"]
    instagram = c["instagram"]
    facebook = c["facebook"]
    builder = None
    status = "success"
    latency = 0.0
    has_viewport = False
    has_ssl = False
    has_seo = False
    
    # 1. Analizar sitio web si existe
    if web:
        web_lower = web.lower()
        is_social = any(x in web_lower for x in ["instagram.com", "facebook.com", "wa.link", "whatsapp.com", "linktr.ee", "youtube.com"])
        if not is_social:
            scraped = analyze_website_socials_and_cms(web)
            status = scraped.get("status", "error")
            builder = scraped.get("builder")
            latency = scraped.get("latency", 0.0)
            has_viewport = scraped.get("has_viewport", False)
            has_ssl = scraped.get("has_ssl", False)
            has_seo = scraped.get("has_seo", False)
            
            if scraped.get("instagram") and not instagram:
                instagram = scraped["instagram"]
            if scraped.get("facebook") and not facebook:
                facebook = scraped["facebook"]
        else:
            status = "social_url"
            has_ssl = web_lower.startswith("https")
            latency = 0.1
            has_viewport = True
            has_seo = True
    else:
        status = "no_url"
        
    # 2. Búsqueda de respaldo en redes si faltan
    if not instagram:
        instagram = fallback_search_socials(c["name"], ciudad, "instagram")
        
    # 3. Búsqueda de respaldo en Facebook si falta
    if not facebook:
        facebook = fallback_search_socials(c["name"], ciudad, "facebook")
        
    # 4. Calcular puntuación determinista
    calidad_score, calidad_motivo = calculate_service_score(
        servicio=servicio,
        website=web,
        rating=c["rating"],
        instagram=instagram,
        facebook=facebook,
        builder=builder,
        status=status,
        has_viewport=has_viewport,
        has_ssl=has_ssl,
        has_seo=has_seo,
        latency=latency
    )
    
    # 5. Generar propuestas
    local_proposal = generate_local_proposal(c["name"], nicho, servicio, estilo_mensaje, ciudad, web, calidad_motivo, instagram, facebook)
    proposal = None
    if api_key:
        proposal = generate_with_gemini(api_key, c["name"], nicho, ciudad, calidad_motivo, servicio, local_proposal, estilo_mensaje)
    if not proposal:
        proposal = local_proposal
        
    # 6. Enlace WhatsApp
    encoded_text = urllib.parse.quote(proposal)
    whatsapp_link = f"https://api.whatsapp.com/send?phone={c['phone_clean']}&text={encoded_text}"
    
    return {
        "name": c["name"],
        "phone_original": c["phone_original"],
        "phone_clean": c["phone_clean"],
        "address": c["address"],
        "message": proposal,
        "whatsapp_link": whatsapp_link,
        "status": "[Pendiente]",
        "website": web,
        "instagram": instagram,
        "facebook": facebook,
        "rating": c["rating"],
        "calidad_score": calidad_score,
        "calidad_motivo": calidad_motivo
    }

@app.route("/api/search", methods=["POST"])
def search_businesses():
    # 0. Validar sesión real del usuario en el backend
    user_session = session.get("user")
    if not user_session:
        return jsonify({
            "status": "error",
            "message": "Por favor, iniciá sesión con Google para poder realizar prospecciones."
        }), 401
        
    email = user_session["email"]
    user_profile = get_user_profile(email)
    
    req_data = request.json or {}
    nicho = req_data.get("nicho", "").strip()
    ciudad = req_data.get("ciudad", "").strip()
    zona = req_data.get("zona", "").strip()
    servicio = req_data.get("servicio", "Desarrollo Web").strip()
    cantidad = req_data.get("cantidad", 20)
    api_key = req_data.get("api_key", "").strip()
    
    if not nicho or not ciudad or not zona or not servicio:
        return jsonify({"status": "error", "message": "Nicho, Ciudad, Zona y Servicio son campos obligatorios."}), 400
        
    # Validar que el servicio ingresado sea un servicio profesional coherente
    if not validate_service_string(servicio):
        return jsonify({
            "status": "error", 
            "message": "Servicio no reconocido. Por favor, introduce un servicio profesional típico (ej: Desarrollo Web, Community Manager, Empleado, Ventas, Asistente, etc.)."
        }), 400

    try:
        cantidad = int(cantidad)
        if cantidad <= 0:
            cantidad = 20
        elif cantidad > 50:
            # Enforzar tope seguro de spam (máximo 50 leads por búsqueda)
            cantidad = 50
    except ValueError:
        cantidad = 20

    # Validar límite de créditos si es plan gratuito
    if user_profile.get("plan") == "free":
        consumed = user_profile.get("leads_consumed", 0)
        if consumed + cantidad > 40:
            return jsonify({
                "status": "error",
                "code": "LIMIT_EXCEEDED",
                "message": f"Esta consulta supera el límite de 40 leads diarios de tu plan gratuito (ya consumiste {consumed} hoy). Pásate a Premium para buscar sin restricciones."
            }), 403


    # Determinar código de área por defecto para la ciudad
    default_area_code = get_area_code_by_city(ciudad)

    # Construir consulta
    query = f"{nicho} en {zona}, {ciudad}, Argentina"
    maps_url = f"https://www.google.com/maps/search/{urllib.parse.quote(query)}"
    
    try:
        # 1. Obtener HTML inicial para extraer el pb
        print(f"Realizando búsqueda inicial en: {maps_url}")
        resp = requests.get(maps_url, headers=HEADERS, timeout=10)
        if resp.status_code != 200:
            return jsonify({"status": "error", "message": "Error al conectar con Google Maps."}), 500
            
        # Buscar el pb parameter en el HTML
        match = re.search(r'href="([^"]*tbm=map[^"]*pb=[^"]*)"', resp.text)
        if not match:
            match = re.search(r'pb=([^"&]*)', resp.text)
            
        if not match:
            return jsonify({"status": "error", "message": "No se pudo iniciar la sesión de búsqueda de Google Maps."}), 500
            
        pb_param = match.group(1)
        if "pb=" in pb_param:
            pb_value = re.search(r'pb=([^&]*)', pb_param).group(1)
        else:
            pb_value = pb_param
            
        # Decodificar el pb para manipular el límite
        pb_decoded = urllib.parse.unquote(pb_value)
        
        # Reemplazar el límite de resultados original (!7i20 por defecto) con la cantidad deseada (!7iXX)
        # La cantidad de resultados solicitada se mapea a !7i{cantidad}
        limit_token = f"!7i{cantidad}"
        if "!7i20" in pb_decoded:
            pb_decoded_modified = pb_decoded.replace("!7i20", limit_token)
        else:
            # Reemplazar cualquier patrón !7i seguido de dígitos
            pb_decoded_modified = re.sub(r'!7i\d+', limit_token, pb_decoded)
            
        # Volver a codificar
        pb_encoded = urllib.parse.quote(pb_decoded_modified)
        
        # 2. Hacer la llamada al API interna de Google Maps (search?tbm=map)
        search_api_url = f"https://www.google.com/search?tbm=map&authuser=0&hl=es&gl=ar&q={urllib.parse.quote(query)}&pb={pb_encoded}"
        print(f"Llamando a la API interna: {search_api_url}")
        tbm_resp = requests.get(search_api_url, headers=HEADERS, timeout=12)
        
        if tbm_resp.status_code != 200 or not tbm_resp.text.startswith(")]}'"):
            return jsonify({"status": "error", "message": "Google Maps denegó la consulta de datos estructurados."}), 500
            
        # Parsear JSON
        clean_text = tbm_resp.text[4:].strip()
        map_data = json.loads(clean_text)
        
        # Comprobar si hay resultados en la posición [64]
        if len(map_data) < 65 or not map_data[64]:
            return jsonify({"status": "success", "data": [], "message": "No se encontraron negocios para esa búsqueda."})
            
        raw_businesses = map_data[64]
        candidates = []
        
        count_valid = 0
        for entry in raw_businesses:
            if count_valid >= cantidad:
                break
                
            if not isinstance(entry, list) or len(entry) < 2:
                continue
            info = entry[1]
            if not isinstance(info, list) or len(info) < 12:
                continue
                
            # Extraer Nombre
            name = info[11] if len(info) > 11 else None
            if not name:
                continue
                
            # Extraer Teléfono
            phone_raw = None
            if len(info) > 178 and info[178] and isinstance(info[178], list) and len(info[178]) > 0:
                phone_sub = info[178][0]
                if isinstance(phone_sub, list) and len(phone_sub) > 0:
                    phone_raw = phone_sub[0]
            
            # REGLA CRUCIAL: Si no tiene teléfono comercial registrado en Google Maps, descártalo inmediatamente
            if not phone_raw:
                continue
                
            # Extraer Dirección / Ubicación
            address = None
            if len(info) > 39 and info[39]:
                address = info[39]
            elif len(info) > 18 and info[18]:
                address = info[18]
            else:
                address = "Dirección no disponible"
                
            # Limpiar Teléfono a formato internacional de WhatsApp
            phone_clean = clean_argentine_phone(phone_raw, default_area_code)
            
            # Extraer URLs recursivamente y clasificar
            all_urls = extract_all_external_urls(info)
            links = get_best_links(all_urls)
            
            rating = None
            if len(info) > 4 and info[4] and isinstance(info[4], list) and len(info[4]) > 7:
                rating = info[4][7]
                
            candidates.append({
                "name": name,
                "phone_original": phone_raw,
                "phone_clean": phone_clean,
                "address": address,
                "rating": rating,
                "website": links.get("website"),
                "instagram": links.get("instagram"),
                "facebook": links.get("facebook"),
                "whatsapp": links.get("whatsapp"),
                "booking": links.get("booking"),
                "other_social": links.get("other_social")
            })
            count_valid += 1


# Continuación de /api/search
        # Procesar todos los candidatos en paralelo para resolver webs, redes sociales, puntuaciones y propuestas
        estilo_mensaje = req_data.get("estilo_mensaje", "argentino").strip()
        extracted_businesses = []
        with ThreadPoolExecutor(max_workers=15) as executor:
            futures = [
                executor.submit(process_candidate_details, c, ciudad, servicio, nicho, api_key, estilo_mensaje)
                for c in candidates
            ]
            extracted_businesses = [f.result() for f in futures]
            
        # Actualizar base de datos de leads consumidos si es gratis
        if user_profile.get("plan") == "free":
            users = load_users()
            if email in users:
                users[email]["leads_consumed"] += len(extracted_businesses)
                save_users(users)
                user_profile = users[email]
                
        return jsonify({
            "status": "success",
            "data": extracted_businesses,
            "leads_consumed_today": user_profile.get("leads_consumed", 15)
        })
        
    except Exception as e:
        import traceback
        error_msg = f"Error procesando búsqueda: {e}\n{traceback.format_exc()}"
        print(error_msg)
        try:
            with open("server_errors.log", "a", encoding="utf-8") as f:
                f.write(f"--- ERROR AT {datetime.datetime.now()} ---\n{error_msg}\n")
        except Exception as log_err:
            print(f"Failed to write to log file: {log_err}")
        return jsonify({"status": "error", "message": f"Ocurrió un error en el servidor: {str(e)}"}), 500

@app.route("/api/export", methods=["POST"])
def export_excel():
    req_data = request.json or {}
    items = req_data.get("items", [])
    
    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospectos B2B"
    
    # Habilitar líneas de cuadrícula
    ws.views.sheetView[0].showGridLines = True
    
    # Estilos
    font_title = Font(name="Segoe UI", size=16, bold=True, color="2C3E50")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)
    font_link = Font(name="Segoe UI", size=10, bold=True, color="0078D4", underline="single")
    
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Título en la primera fila
    ws.merge_cells("A1:I1")
    ws["A1"] = "PROSPECCIÓN B2B - PROPUESTAS DE DESARROLLO Y SERVICIOS"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Dejar una fila vacía
    ws.row_dimensions[2].height = 15
    
    # Encabezados (9 columnas)
    headers = [
        "Nombre del Negocio (Real)",
        "Teléfono Original (Google Maps)",
        "Sitio Web Encontrado",
        "Instagram Detectado",
        "Calidad (1-100)",
        "Dirección / Ubicación",
        "Mensaje Generado por IA",
        "Enlace de WhatsApp",
        "Estado"
    ]
    
    for col_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    ws.row_dimensions[3].height = 28
    
    # Llenar datos
    for idx, item in enumerate(items, 4):
        ws.row_dimensions[idx].height = 60 # Espacio para el wrap_text
        
        # Columna A: Nombre
        cell_name = ws.cell(row=idx, column=1, value=item.get("name", ""))
        cell_name.font = font_data
        cell_name.alignment = align_left
        cell_name.border = thin_border
        
        # Columna B: Teléfono Original
        cell_phone = ws.cell(row=idx, column=2, value=item.get("phone_original", ""))
        cell_phone.font = font_data
        cell_phone.alignment = align_center
        cell_phone.border = thin_border
        
        # Columna C: Sitio Web Encontrado
        web_url = item.get("website", "")
        if web_url:
            web_formula = f'=HYPERLINK("{web_url}", "{web_url[:40]}...")'
            cell_web = ws.cell(row=idx, column=3, value=web_formula)
            cell_web.font = font_link
        else:
            cell_web = ws.cell(row=idx, column=3, value="No registrado")
            cell_web.font = font_data
        cell_web.alignment = align_left
        cell_web.border = thin_border
        
        # Columna D: Instagram Detectado
        ig_url = item.get("instagram", "")
        if ig_url:
            ig_formula = f'=HYPERLINK("{ig_url}", "{ig_url[:40]}...")'
            cell_ig = ws.cell(row=idx, column=4, value=ig_formula)
            cell_ig.font = font_link
        else:
            cell_ig = ws.cell(row=idx, column=4, value="No registrado")
            cell_ig.font = font_data
        cell_ig.alignment = align_left
        cell_ig.border = thin_border
        
        # Columna E: Calidad (1-100)
        calidad_val = item.get("calidad_score", "")
        calidad_mot = item.get("calidad_motivo", "")
        calidad_text = f"{calidad_val} ({calidad_mot})" if calidad_val else "N/A"
        cell_cal = ws.cell(row=idx, column=5, value=calidad_text)
        cell_cal.font = font_data
        cell_cal.alignment = align_center
        cell_cal.border = thin_border
        
        # Columna F: Dirección
        cell_addr = ws.cell(row=idx, column=6, value=item.get("address", ""))
        cell_addr.font = font_data
        cell_addr.alignment = align_left
        cell_addr.border = thin_border
        
        # Columna G: Mensaje
        cell_msg = ws.cell(row=idx, column=7, value=item.get("message", ""))
        cell_msg.font = font_data
        cell_msg.alignment = align_left
        cell_msg.border = thin_border
        
        # Columna H: Enlace de WhatsApp (=HYPERLINK)
        link_url = item.get("whatsapp_link", "")
        formula = f'=HYPERLINK("{link_url}", "📲 Enviar Propuesta")'
        cell_link = ws.cell(row=idx, column=8, value=formula)
        cell_link.font = font_link
        cell_link.alignment = align_center
        cell_link.border = thin_border
        
        # Columna I: Estado
        cell_status = ws.cell(row=idx, column=9, value=item.get("status", "[Pendiente]"))
        cell_status.font = font_data
        cell_status.alignment = align_center
        cell_status.border = thin_border
        
        # Cebra
        if idx % 2 == 0:
            for col in range(1, 10):
                if col not in [3, 4, 8] or (col == 3 and not web_url) or (col == 4 and not ig_url): # No pisar color de enlaces activos
                    ws.cell(row=idx, column=col).fill = fill_zebra
                    
    # Autoajustar columnas
    column_widths = {
        "A": 30, # Nombre
        "B": 20, # Teléfono
        "C": 30, # Sitio Web
        "D": 30, # Instagram
        "E": 18, # Calidad
        "F": 35, # Dirección
        "G": 50, # Mensaje
        "H": 22, # Enlace WhatsApp
        "I": 15  # Estado
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    # Guardar en memoria
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    
    return send_file(
        excel_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Prospectos_B2B.xlsx"
    )

if __name__ == "__main__":
    app.run(debug=True, host='0.0.0.0', port=5000)
